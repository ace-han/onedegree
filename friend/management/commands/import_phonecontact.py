'''
Created on Sep 6, 2015

@author: ace
'''
import argparse
import re

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from account.models import Profile
from friend.models import PhoneContactRecord


class Command(BaseCommand):
    help = 'Batch import Phone Contact from excel file'

    def add_arguments(self, parser):
        parser.add_argument('filepath',
                            type=argparse.FileType('rb'),
                            help='the filepath on the system to import')

    def handle(self, *args, **options):
        wb = load_workbook(options['filepath'], use_iterators = True)
        ws = wb.worksheets[0]
        row_counter = 1
        newly_created_counter, updated_counter, err_rows = 0, 0, 0
        '''
        from_phone_num    to_phone_num(comma separated)
        1234567890        1234567891, 1234567892, 1234567893, 1234567894
        1234567891        1234567895, 1234567896, 1234567897, 1234567898
        1234567892        1234567895, 1234567896, 1234567897, 1234567898
        '''
        phone_num_profile_dict = dict( (profile.phone_num, profile) for profile in Profile.objects.all())
        for row in ws.iter_rows(row_offset=1):
            row_counter += 1
            from_phone_num = str(row[0].value or '')
            from_profile = phone_num_profile_dict.get(from_phone_num)
            if not from_profile:
                self.stderr.write('Line: %d, no profile for from_phone_num: %s' % (row_counter, from_phone_num,) )
                err_rows += 1
                continue
            
            # actually in this style means always newly_created
            PhoneContactRecord.objects.filter(from_profile=from_profile).delete()
            
            to_phone_num_strs = re.split('[,锛� ]+', str(row[1].value or '') )
            to_profiles = [ phone_num_profile_dict[to_phone_num] for to_phone_num in to_phone_num_strs if to_phone_num in phone_num_profile_dict ]
            for to_profile in to_profiles:
                defaults = {
                    'to_phone_num': to_profile.phone_num
                }
                try:
                    contact_record, newly_created = PhoneContactRecord.objects.get_or_create(from_profile=from_profile, 
                                                                                 to_profile=to_profile, defaults=defaults)
                except Exception as e:
                    raise CommandError('Line: %d encounter error: %s' % (row_counter, e, ) )
                else:
                    if newly_created:
                        # actually in this style means always newly_created
                        newly_created_counter += 1
                    else:
                        for key, value in defaults.items():
                            setattr(contact_record, key, value)
                        contact_record.save()
                        updated_counter += 1
        self.stdout.write('newly created: %d, updated: %d, total: %d, rows: %d, err_rows: %d' % (newly_created_counter, 
                                                                         updated_counter, 
                                                                         newly_created_counter+updated_counter, 
                                                                         row_counter, 
                                                                         err_rows, ) )
            
