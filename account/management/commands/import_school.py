'''
Created on Sep 6, 2015

@author: ace
'''
import argparse

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from account.models import School


class Command(BaseCommand):
    help = 'Batch import school from excel file'

    def add_arguments(self, parser):
        parser.add_argument('filepath',
                            type=argparse.FileType('rb'),
                            help='the filepath on the system to import')

    def handle(self, *args, **options):
        wb = load_workbook(options['filepath'], use_iterators = True)
        ws = wb.worksheets[0]
        row_counter = 1
        newly_created_counter, updated_counter = 0, 0
        '''
        Line_0 name    type            remark
        Line_1 A中学    high_school    A中学
        Line_2 B大学    college        B大学
        '''
        for row in ws.iter_rows(row_offset=1):
            row_counter += 1
            name = row[0].value
            
            defaults = {
                'type': row[1].value,
                'remark': row[2].value or '',
            }
            try:
                school, newly_created = School.objects.get_or_create(name=name, defaults=defaults)
            except Exception as e:
                raise CommandError('Line: %d encounter error: %s' % (row_counter, e, ) )
            else:
                if newly_created:
                    newly_created_counter += 1
                else:
                    for key, value in defaults.items():
                        setattr(school, key, value)
                    school.save()
                    updated_counter += 1
        self.stdout.write('newly created: %d, updated: %d, total: %d' % (newly_created_counter, 
                                                                         updated_counter, 
                                                                         newly_created_counter+updated_counter, ) )
            
