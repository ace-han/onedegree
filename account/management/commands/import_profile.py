'''
Created on Sep 6, 2015

@author: ace
'''
import argparse
from functools import partial
import re

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from taggit.models import Tag

from account.models import Profile, School
from tag.models import TreeTag


class Command(BaseCommand):
    help = 'Batch import Profile from excel file'

    def add_arguments(self, parser):
        parser.add_argument('filepath',
                            type=argparse.FileType('rb'),
                            help='the filepath on the system to import')

    def handle(self, *args, **options):
        wb = load_workbook(options['filepath'], use_iterators = True)
        ws = wb.worksheets[0]
        row_counter = 1
        newly_created_counter, updated_counter = 0, 0
        '''
        phone_num    gender    city        whatsup    high_school    college         occupations(3 tops)               tags(comma separated)
        1234567890    male    beijing      aaa            A中学            A大学        occupations1,occupations2        child1,child2, test
        1234567891            shanghai     bbb            B中学            B大学        occupations                        child11, child21, test11
        1234567892    female  guangzhou    ccc            C中学            C大学        child22, test21                  child22, test21
        '''
        name_occupation_dict = dict( (tag.name, tag) for tag in TreeTag.objects.get(slug='occupation-tag-root').get_descendants(include_self=False))
        name_tag_dict = dict( (tag.name, tag) for tag in Tag.objects.all() )
        name_high_school_dict = dict( (school.name, school) for school in School.objects.filter(type='high_school') )
        name_college_dict = dict( (school.name, school) for school in School.objects.filter(type='college') )
        for row in ws.iter_rows(row_offset=1):
            row_counter += 1
            phone_num = row[0].value or ''
            gender_str = (row[1].value or '').strip()
            if gender_str:
                gender = 1 if gender_str.lower()=='male' else 0
            else:
                gender = None 
            defaults = {
                'gender': gender,
                'city': row[2].value or '',
                'whatsup': row[3].value or '',
            }
            if row[4].value:
                defaults['high_school'] = name_high_school_dict.get(row[4].value.strip())
            if row[5].value:
                defaults['college'] = name_college_dict.get(row[5].value.strip())
            occupations_strs = re.split('[,锛� ]+', row[6].value or '')
            occupations = [ name_occupation_dict[tag_name] for tag_name in occupations_strs if tag_name in name_occupation_dict ]
            tag_strs = re.split('[,锛� ]+', row[7].value or '')
            tags = []
            for tag_name in tag_strs:
                if tag_name in name_tag_dict:
                    tags.append(name_tag_dict[tag_name])
                else:
                    if not tag_name:
                        continue
                    tag = Tag(name=tag_name)
                    # take advantage of Unihandecoder
                    tag.slugify = partial(TreeTag.slugify, tag)   
                    tag.save()
                    tags.append(tag)
            print('phone_num', phone_num, 'defaults', defaults)
            try:
                profile, newly_created = Profile.objects.get_or_create(phone_num=phone_num, defaults=defaults)
            except Exception as e:
                raise CommandError('Line: %d encounter error: %s' % (row_counter, e, ) )
            else:
                if newly_created:
                    newly_created_counter += 1
                else:
                    for key, value in defaults.items():
                        setattr(profile, key, value)
                    profile.save()
                    updated_counter += 1
                if occupations:
                    profile.occupations.set(*occupations)
                if tag_strs:
                    # if the tag is not in the system it would create a new one
                    profile.tags.set(*tags)
        self.stdout.write('newly created: %d, updated: %d, total: %d' % (newly_created_counter, 
                                                                         updated_counter, 
                                                                         newly_created_counter+updated_counter, ) )
            
